"""
r: read only
w: write only
a: append
r+: read and write with overwrite existing file, not create new file
w+: read and write with clear all content of existing file, create new file if not exist
a+: read and write with append content to the last of existing file, create new file if not exist
x: create new file, bao loi neu file da ton tai
t: text mode
b: binary mode
"""
""" read file """
import os
from pathlib import Path
import openpyxl as oxl
# file_obj  = open("my_file.txt", "rt", encoding="utf-8")
# content = file_obj.read()
# print(content)
# print(file_obj.mode)
# print(file_obj.name)

# file_obj.close() 
# os.rename("my_file.txt", "my_file_1.txt")
# os.remove("my_file_1.txt")
# print(file_obj.closed)

# Tao, Xoa thu muc
# os.mkdir("my_folder")
# os.rmdir("my_folder")

# with open("my_file.txt", "rt", encoding="utf-8") as file_obj:
    # doc toan bo noi dung => read()
    # content = file_obj.read(100)
    # content = file_obj.read()
    # print(content)

    # current_poiner = file_obj.tell()
    # print("Current Pointer:", current_poiner)

    # dua con tro ve vi tri 30
    # seek_obj = file_obj.seek(30)
    
    # tra ve vi tri con tro hien tai
    # current_poiner = file_obj.tell()
    # print("Current Pointer after seek:", current_poiner)

    # doc 1 dong
    # read_line = file_obj.readline()
    # print(read_line)
    # read_line = file_obj.readline()
    # print(read_line)
    # read_line = file_obj.readline()
    # print(read_line)

    # doc nhieu dong
    # read_lines = file_obj.readlines()
    # for read_line in read_lines:
    #     print(read_line)

""" write file """
# with open("my_file.txt", "rt+", encoding="utf-8") as f:
#     f.write("ghi de, khong tao file moi neu file chua ton tai\n")

# with open("my_new_file.txt", "wt", encoding="utf-8") as f:
#     f.write("Tao file moi, xoa noi dung file hien tai neu file dang ton tai\n")

""" append file """
# with open("my_new_file123.txt", "at", encoding="utf-8") as f:
#     f.write("Noi dung nay duoc dinh kem xuong cuoi cung neu file dang ton tai\n")

""" create new file """
# with open("my_new_file.txt", "x", encoding="utf-8") as f:
#     f.write("Noi dung moi nay duoc tao file moi\n")

# Kiem tra thu muc ton tai
# path = Path("my_module")
# print(path.exists())

# Tao, Xoa thu muc
#path.mkdir()
#path.rmdir()

# Liet ke tat ca cac file trong thu muc:
# path = Path() # tro ve thu muc hien tai
# for file in path.glob('*.py'):
#     print(file)

wb = oxl.load_workbook("price_list.xlsx")
sheet = wb['Sheet1']
# cell1 = sheet['a1']
# cell2 = sheet.cell(2,2)
# print(cell1.value)
# print(cell2.value)

# print(sheet.max_row)
for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    cell.value *= 1.1

wb.save("price_list.xlsx")